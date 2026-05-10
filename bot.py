import logging
import os
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from reportlab.lib.pagesizes import A4
from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
from reportlab.lib.units import cm
from reportlab.lib import colors
from reportlab.platypus import SimpleDocTemplate, Paragraph, Spacer, Table, TableStyle
from reportlab.pdfbase import pdfmetrics
from reportlab.pdfbase.ttfonts import TTFont
from telegram import Update, ReplyKeyboardMarkup, ReplyKeyboardRemove
from telegram.ext import (
    Application, CommandHandler, MessageHandler,
    ConversationHandler, ContextTypes, filters
)
from datetime import datetime
import re

logging.basicConfig(level=logging.INFO)
logger = logging.getLogger(__name__)

# ── Шаги анкеты ──────────────────────────────────────────────
(
    FIO, AGE, CITY, PHONE, EMAIL,
    EDUCATION, SPECIALTY, EXP_YEARS, EXP_COMPANIES, EXP_ROLES,
    ACHIEVEMENTS, SALES_SKILLS, TOOLS, KPI, DEAL_SIZE,
    CLIENTS, PRODUCT, MOTIVATION, SALARY, SCHEDULE,
    ENGLISH, DRIVING, READINESS_TRAVEL, REMOTE,
    LINKEDIN, PORTFOLIO, STRENGTHS, WEAKNESSES,
    ABOUT, QUESTIONS
) = range(30)

QUESTIONS = {
    FIO:               "👤 *1/30* — Ваше полное ФИО?",
    AGE:               "🎂 *2/30* — Ваш возраст?",
    CITY:              "🏙 *3/30* — В каком городе вы находитесь?",
    PHONE:             "📞 *4/30* — Ваш номер телефона?",
    EMAIL:             "📧 *5/30* — Ваш email?",
    EDUCATION:         "🎓 *6/30* — Образование (ВУЗ, специальность, год окончания)?",
    SPECIALTY:         "📚 *7/30* — Дополнительные курсы или сертификаты в сфере продаж/маркетинга?",
    EXP_YEARS:         "📅 *8/30* — Сколько лет общего опыта в продажах/маркетинге?",
    EXP_COMPANIES:     "🏢 *9/30* — Перечислите последние 3 места работы (компания и период)?",
    EXP_ROLES:         "💼 *10/30* — Какие должности вы занимали?",
    ACHIEVEMENTS:      "🏆 *11/30* — Ваши главные достижения в продажах (конкретные цифры)?",
    SALES_SKILLS:      "🎯 *12/30* — Какие техники продаж вы знаете и применяете?",
    TOOLS:             "🛠 *13/30* — Какие CRM-системы и маркетинговые инструменты вы используете?",
    KPI:               "📊 *14/30* — Какие KPI вы выполняли на последнем месте работы?",
    DEAL_SIZE:         "💰 *15/30* — Средний чек сделки, с которой вы работали?",
    CLIENTS:           "🤝 *16/30* — Какие типы клиентов: B2B, B2C или оба?",
    PRODUCT:           "📦 *17/30* — Что именно вы продавали (товары/услуги/digital)?",
    MOTIVATION:        "💡 *18/30* — Почему вас интересует именно наша вакансия?",
    SALARY:            "💵 *19/30* — Ваши ожидания по зарплате (фикс + бонус)?",
    SCHEDULE:          "🕐 *20/30* — Предпочтительный формат работы?",
    ENGLISH:           "🌍 *21/30* — Уровень английского языка?",
    DRIVING:           "🚗 *22/30* — Есть ли у вас водительские права и автомобиль?",
    READINESS_TRAVEL:  "✈️ *23/30* — Готовы ли вы к командировкам?",
    REMOTE:            "🏠 *24/30* — Рассматриваете ли вы удалённую работу?",
    LINKEDIN:          "🔗 *25/30* — Ссылка на ваш LinkedIn или HH.ru профиль (или напишите «нет»)?",
    PORTFOLIO:         "📁 *26/30* — Есть ли портфолио или кейсы? Укажите ссылку или опишите.",
    STRENGTHS:         "⭐️ *27/30* — Ваши 3 главных профессиональных сильных стороны?",
    WEAKNESSES:        "🔍 *28/30* — Над чем вы сейчас работаете / что хотите улучшить в себе?",
    ABOUT:             "📝 *29/30* — Расскажите коротко о себе (2–3 предложения)?",
    30-1:              "❓ *30/30* — Есть ли у вас вопросы к нашей компании или вакансии?",
}

QUESTION_KEYS = [
    "ФИО", "Возраст", "Город", "Телефон", "Email",
    "Образование", "Доп. курсы", "Опыт (лет)", "Места работы", "Должности",
    "Достижения", "Техники продаж", "CRM/Инструменты", "KPI", "Средний чек",
    "Тип клиентов", "Что продавал(а)", "Мотивация", "Ожидания по ЗП", "График",
    "Английский", "Водительские права", "Командировки", "Удалёнка",
    "LinkedIn/HH", "Портфолио", "Сильные стороны", "Зоны роста",
    "О себе", "Вопросы к компании"
]

KEYBOARD_SCHEDULE   = [["Офис", "Удалёнка", "Гибрид"]]
KEYBOARD_ENGLISH    = [["Нет", "Базовый", "Intermediate", "Upper-Intermediate", "Fluent"]]
KEYBOARD_DRIVING    = [["Да, есть права и авто", "Есть права, нет авто", "Нет"]]
KEYBOARD_YESNO      = [["Да", "Нет"]]
KEYBOARD_CLIENTS    = [["B2B", "B2C", "Оба"]]

EXCEL_FILE = "резюме_кандидатов.xlsx"
ADMIN_CHAT_ID = int(os.getenv("ADMIN_CHAT_ID", "0"))


# ── Хелперы Excel ─────────────────────────────────────────────
def init_excel():
    if os.path.exists(EXCEL_FILE):
        return
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Кандидаты"
    header_fill = PatternFill("solid", fgColor="1F3864")
    header_font = Font(bold=True, color="FFFFFF", size=11)
    headers = ["№", "Дата"] + QUESTION_KEYS
    for col, h in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=h)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center", wrap_text=True)
    ws.row_dimensions[1].height = 30
    ws.column_dimensions["A"].width = 5
    ws.column_dimensions["B"].width = 16
    for i in range(3, len(headers) + 1):
        ws.column_dimensions[openpyxl.utils.get_column_letter(i)].width = 28
    wb.save(EXCEL_FILE)

def save_to_excel(answers: list):
    init_excel()
    wb = openpyxl.load_workbook(EXCEL_FILE)
    ws = wb.active
    row_num = ws.max_row + 1
    alt_fill = PatternFill("solid", fgColor="EBF0FA")
    border = Border(
        left=Side(style="thin", color="CCCCCC"),
        right=Side(style="thin", color="CCCCCC"),
        top=Side(style="thin", color="CCCCCC"),
        bottom=Side(style="thin", color="CCCCCC"),
    )
    row = [row_num - 1, datetime.now().strftime("%d.%m.%Y %H:%M")] + answers
    for col, val in enumerate(row, 1):
        cell = ws.cell(row=row_num, column=col, value=val)
        cell.alignment = Alignment(wrap_text=True, vertical="top")
        cell.border = border
        if row_num % 2 == 0:
            cell.fill = alt_fill
    ws.row_dimensions[row_num].height = 60
    wb.save(EXCEL_FILE)


# ── Хелпер PDF ────────────────────────────────────────────────
def generate_pdf(answers: list, username: str) -> str:
    filename = f"резюме_{username}_{datetime.now().strftime('%Y%m%d_%H%M%S')}.pdf"
    doc = SimpleDocTemplate(filename, pagesize=A4,
                            leftMargin=2*cm, rightMargin=2*cm,
                            topMargin=2*cm, bottomMargin=2*cm)
    styles = getSampleStyleSheet()
    title_style = ParagraphStyle("title", parent=styles["Title"],
                                 fontSize=18, textColor=colors.HexColor("#1F3864"),
                                 spaceAfter=6)
    sub_style = ParagraphStyle("sub", parent=styles["Normal"],
                               fontSize=10, textColor=colors.grey, spaceAfter=16)
    section_style = ParagraphStyle("section", parent=styles["Normal"],
                                   fontSize=11, textColor=colors.HexColor("#1F3864"),
                                   fontName="Helvetica-Bold", spaceAfter=4, spaceBefore=10)
    answer_style = ParagraphStyle("ans", parent=styles["Normal"],
                                  fontSize=10, leading=14, spaceAfter=2)

    story = []
    story.append(Paragraph("📋 Анкета кандидата", title_style))
    story.append(Paragraph(f"Вакансия: Продажи / Маркетинг  •  Дата: {datetime.now().strftime('%d.%m.%Y')}", sub_style))

    table_data = []
    for i, (key, val) in enumerate(zip(QUESTION_KEYS, answers)):
        q_cell = Paragraph(f"<b>{key}</b>", answer_style)
        a_cell = Paragraph(str(val) if val else "—", answer_style)
        table_data.append([q_cell, a_cell])

    table = Table(table_data, colWidths=[5*cm, 12*cm])
    table.setStyle(TableStyle([
        ("BACKGROUND", (0, 0), (0, -1), colors.HexColor("#EBF0FA")),
        ("VALIGN", (0, 0), (-1, -1), "TOP"),
        ("GRID", (0, 0), (-1, -1), 0.5, colors.HexColor("#CCCCCC")),
        ("ROWBACKGROUNDS", (0, 0), (-1, -1), [colors.white, colors.HexColor("#F7F9FD")]),
        ("TOPPADDING", (0, 0), (-1, -1), 6),
        ("BOTTOMPADDING", (0, 0), (-1, -1), 6),
        ("LEFTPADDING", (0, 0), (-1, -1), 8),
        ("RIGHTPADDING", (0, 0), (-1, -1), 8),
    ]))
    story.append(table)
    doc.build(story)
    return filename


# ── Handlers ──────────────────────────────────────────────────
async def start(update: Update, context: ContextTypes.DEFAULT_TYPE):
    context.user_data.clear()
    context.user_data["answers"] = []
    await update.message.reply_text(
        "👋 Добро пожаловать!\n\n"
        "Это анкета для кандидатов на вакансию *Продажи / Маркетинг*.\n\n"
        "📋 Вас ждут *30 вопросов* — отвечайте честно и развёрнуто.\n"
        "⏱ Займёт около 10–15 минут.\n\n"
        "Начнём? 🚀",
        parse_mode="Markdown",
        reply_markup=ReplyKeyboardMarkup([["✅ Начать анкету"]], resize_keyboard=True)
    )
    return FIO

async def ask_fio(update: Update, context: ContextTypes.DEFAULT_TYPE):
    await update.message.reply_text(
        QUESTIONS[FIO], parse_mode="Markdown",
        reply_markup=ReplyKeyboardRemove()
    )
    return FIO

def make_handler(current_step, next_step, next_question, keyboard=None):
    async def handler(update: Update, context: ContextTypes.DEFAULT_TYPE):
        context.user_data["answers"].append(update.message.text)
        kb = ReplyKeyboardMarkup(keyboard, resize_keyboard=True, one_time_keyboard=True) if keyboard else ReplyKeyboardRemove()
        await update.message.reply_text(next_question, parse_mode="Markdown", reply_markup=kb)
        return next_step
    return handler

# Последний вопрос → сохранение
async def finish(update: Update, context: ContextTypes.DEFAULT_TYPE):
    context.user_data["answers"].append(update.message.text)
    answers = context.user_data["answers"]
    user = update.effective_user
    username = user.username or user.first_name or str(user.id)

    await update.message.reply_text(
        "⏳ Сохраняю вашу анкету...", reply_markup=ReplyKeyboardRemove()
    )

    # Excel
    save_to_excel(answers)

    # PDF
    pdf_file = generate_pdf(answers, username)

    await update.message.reply_text(
        "✅ *Анкета принята!*\n\n"
        "Спасибо, что уделили время. Мы свяжемся с вами в ближайшие 2–3 рабочих дня.\n\n"
        "Удачи! 🍀",
        parse_mode="Markdown"
    )

    # Отправить HR
    if ADMIN_CHAT_ID:
        name = answers[0] if answers else username
        await context.bot.send_message(
            ADMIN_CHAT_ID,
            f"📥 *Новая анкета от кандидата!*\n\n"
            f"👤 Имя: {name}\n"
            f"📱 Telegram: @{username}\n"
            f"📅 Дата: {datetime.now().strftime('%d.%m.%Y %H:%M')}",
            parse_mode="Markdown"
        )
        with open(pdf_file, "rb") as f:
            await context.bot.send_document(ADMIN_CHAT_ID, f, filename=f"Анкета_{name}.pdf")

    os.remove(pdf_file)
    return ConversationHandler.END

async def cancel(update: Update, context: ContextTypes.DEFAULT_TYPE):
    await update.message.reply_text(
        "❌ Анкета отменена. Напишите /start чтобы начать заново.",
        reply_markup=ReplyKeyboardRemove()
    )
    return ConversationHandler.END


# ── Построение цепочки вопросов ───────────────────────────────
def build_conversation():
    steps = [
        (FIO,              AGE,             QUESTIONS[AGE],              None),
        (AGE,              CITY,            QUESTIONS[CITY],             None),
        (CITY,             PHONE,           QUESTIONS[PHONE],            None),
        (PHONE,            EMAIL,           QUESTIONS[EMAIL],            None),
        (EMAIL,            EDUCATION,       QUESTIONS[EDUCATION],        None),
        (EDUCATION,        SPECIALTY,       QUESTIONS[SPECIALTY],        None),
        (SPECIALTY,        EXP_YEARS,       QUESTIONS[EXP_YEARS],        None),
        (EXP_YEARS,        EXP_COMPANIES,   QUESTIONS[EXP_COMPANIES],    None),
        (EXP_COMPANIES,    EXP_ROLES,       QUESTIONS[EXP_ROLES],        None),
        (EXP_ROLES,        ACHIEVEMENTS,    QUESTIONS[ACHIEVEMENTS],     None),
        (ACHIEVEMENTS,     SALES_SKILLS,    QUESTIONS[SALES_SKILLS],     None),
        (SALES_SKILLS,     TOOLS,           QUESTIONS[TOOLS],            None),
        (TOOLS,            KPI,             QUESTIONS[KPI],              None),
        (KPI,              DEAL_SIZE,       QUESTIONS[DEAL_SIZE],        None),
        (DEAL_SIZE,        CLIENTS,         QUESTIONS[CLIENTS],          KEYBOARD_CLIENTS),
        (CLIENTS,          PRODUCT,         QUESTIONS[PRODUCT],          None),
        (PRODUCT,          MOTIVATION,      QUESTIONS[MOTIVATION],       None),
        (MOTIVATION,       SALARY,          QUESTIONS[SALARY],           None),
        (SALARY,           SCHEDULE,        QUESTIONS[SCHEDULE],         KEYBOARD_SCHEDULE),
        (SCHEDULE,         ENGLISH,         QUESTIONS[ENGLISH],          KEYBOARD_ENGLISH),
        (ENGLISH,          DRIVING,         QUESTIONS[DRIVING],          KEYBOARD_DRIVING),
        (DRIVING,          READINESS_TRAVEL,QUESTIONS[READINESS_TRAVEL], KEYBOARD_YESNO),
        (READINESS_TRAVEL, REMOTE,          QUESTIONS[REMOTE],           KEYBOARD_YESNO),
        (REMOTE,           LINKEDIN,        QUESTIONS[LINKEDIN],         None),
        (LINKEDIN,         PORTFOLIO,       QUESTIONS[PORTFOLIO],        None),
        (PORTFOLIO,        STRENGTHS,       QUESTIONS[STRENGTHS],        None),
        (STRENGTHS,        WEAKNESSES,      QUESTIONS[WEAKNESSES],       None),
        (WEAKNESSES,       ABOUT,           QUESTIONS[ABOUT],            None),
        (ABOUT,            29,              "❓ *30/30* — Есть ли у вас вопросы к нашей компании или вакансии?", None),
    ]

    handlers = {}
    for curr, nxt, q, kb in steps:
        handlers[curr] = [MessageHandler(filters.TEXT & ~filters.COMMAND,
                                         make_handler(curr, nxt, q, kb))]

    # Последний шаг
    handlers[29] = [MessageHandler(filters.TEXT & ~filters.COMMAND, finish)]

    return handlers


def main():
    token = os.getenv("BOT_TOKEN")
    if not token:
        raise ValueError("Не задан BOT_TOKEN! Добавьте его в переменные окружения.")

    app = Application.builder().token(token).build()

    step_handlers = build_conversation()

    conv = ConversationHandler(
        entry_points=[CommandHandler("start", ask_fio)],
        states=step_handlers,
        fallbacks=[CommandHandler("cancel", cancel)],
    )

    app.add_handler(conv)
    init_excel()
    logger.info("Бот запущен ✅")
    app.run_polling()


if __name__ == "__main__":
    main()
